# -*- coding: utf-8 -*-
# Regenere cartes.js a partir de data/cartes_maitre.xlsx (feuilles "Cartes" et "Liens").
# C'est la SEULE maniere dont cartes.js doit changer : ne jamais l'editer a la main,
# il est ecrase a chaque execution de ce script.
#
# Usage : modifie data/cartes_maitre.xlsx dans Excel/LibreOffice, puis :
#   python generer_cartes.py
import json
import os
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, 'data', 'cartes_maitre.xlsx')
OUT = os.path.join(BASE, 'cartes.js')

# Emoji de secours si la colonne "image" est vide, indexes par id (1..137)
EMOJIS = {
    1: "🦕", 2: "☄️", 3: "🪨", 4: "🎨", 5: "🐑", 6: "🛞", 7: "🐂", 8: "☀️", 9: "📐", 10: "🏹",
    11: "🏗️", 12: "🧱", 13: "⚔️", 14: "👑", 15: "🏛️", 16: "🌋", 17: "♟️", 18: "🛡️", 19: "🍴", 20: "🤟",
    21: "✝️", 22: "💣", 23: "🏙️", 24: "☠️", 25: "🖨️", 26: "⛵", 27: "🪥", 28: "⌚", 29: "🌍", 30: "🔭",
    31: "🌡️", 32: "📖", 33: "🕌", 34: "🔥", 35: "🥐", 36: "🌋", 37: "🧭", 38: "🌡️", 39: "🧺", 40: "🇺🇸",
    41: "🎈", 42: "🏰", 43: "✏️", 44: "🔋", 45: "👑", 46: "🚂", 47: "🥫", 48: "🧴", 49: "🏝️", 50: "🔥",
    51: "🕯️", 52: "📡", 53: "⛏️", 54: "⛓️", 55: "🔧", 56: "🇮🇹", 57: "🎭", 58: "📜", 59: "🤠", 60: "👖",
    61: "☎️", 62: "🎶", 63: "💡", 64: "🏢", 65: "🗼", 66: "🪥", 67: "🥣", 68: "📻", 69: "🎬", 70: "🛩️",
    71: "🧸", 72: "✈️", 73: "🏅", 74: "🍵", 75: "🛥️", 76: "🚢", 77: "⚔️", 78: "🔫", 79: "🚢", 80: "🕊️",
    81: "🖋️", 82: "⚱️", 83: "✈️", 84: "🧫", 85: "🎸", 86: "📉", 87: "⚽", 88: "💣", 89: "🪖", 90: "💥",
    91: "✌️", 92: "🌍", 93: "🏥", 94: "☢️", 95: "🇮🇳", 96: "💨", 97: "🔌", 98: "🛸", 99: "💿", 100: "🏔️",
    101: "🧬", 102: "🚌", 103: "🛰️", 104: "🥇", 105: "🚀", 106: "🧱", 107: "☢️", 108: "🔫", 109: "✊", 110: "🌕",
    111: "🌐", 112: "🎮", 113: "🔢", 114: "🎥", 115: "🎵", 116: "🕹️", 117: "🧱", 118: "🕸️", 119: "✊", 120: "🚩",
    121: "💽", 122: "🗳️", 123: "🐑", 124: "💊", 125: "🏢", 126: "📚", 127: "📘", 128: "📱", 129: "🗳️", 130: "🏃",
    131: "🍕", 132: "🪂", 133: "⚛️", 134: "🧊", 135: "📺", 136: "👽", 137: "🤖",
    # ids 247+ : emojis choisis specifiquement pour chaque carte (en attendant
    # une illustration), plutot que le repli generique par famille.
    247: "💉", 248: "🦠", 249: "🔬", 250: "🔭", 251: "🔌", 252: "🧮", 253: "🛰️", 254: "🔋", 255: "📜", 256: "🪐",
    257: "🐕", 258: "🥜", 259: "✏️", 260: "🏷️", 261: "👶", 262: "🔲", 263: "⌚", 264: "📠", 265: "🖨️", 266: "📼",
    267: "🧮", 268: "💿", 269: "💿", 270: "🎧", 271: "🎬", 272: "🔋", 273: "💡", 274: "🥤", 275: "🧊", 276: "👥",
    277: "🐾", 278: "🌊", 279: "🥵", 280: "🐚", 281: "☄️", 282: "🌋", 283: "🗺️", 284: "☀️", 285: "🔴", 286: "🌌",
    287: "⏳", 288: "🦣", 289: "🌱", 290: "⚫", 291: "🕳️", 292: "🌑", 293: "🦠", 294: "🫧", 295: "🧫", 296: "🦐",
    297: "🐟", 298: "🌿", 299: "🐸", 300: "🥚", 301: "🦴", 302: "💀", 303: "🐭", 304: "🔥", 305: "🪨", 306: "❄️",
    307: "🌍", 308: "🏕️", 309: "🚶", 310: "🐚", 311: "🛶", 312: "🤝", 313: "🎨", 314: "🕯️", 315: "🗿", 316: "🖼️",
    317: "🏹", 318: "🥶", 319: "🐕", 320: "🌡️", 321: "🌾", 322: "🍞",
}

FAMILLE_SLUG = {
    "Histoire": "histoire", "Science": "science", "Invention": "inventions", "Inventions": "inventions",
    "Culture": "culture", "Architecture": "architecture", "Nature": "nature", "Guerre": "guerre",
    "Exploration": "exploration", "Mythologie": "mythologie", "Sport": "sport",
}

# Emoji par defaut pour une carte SANS emoji dedie ET sans image : mieux que "?" partout.
# Cle = famille normalisee (categorie avant le "/", ou famille speciale ci-dessous).
# Cette liste est aussi la liste des categories filtrables (script.js/multi.js
# la reproduisent cote client pour construire les cases a cocher des filtres).
EMOJI_PAR_FAMILLE = {
    "histoire": "📜", "science": "🔬", "inventions": "⚙️", "culture": "🎭",
    "cinema": "🎬", "television": "📺", "jeuxvideo": "🎮",
    "architecture": "🏛️", "nature": "🌿", "guerre": "⚔️", "exploration": "🧭",
    "mythologie": "🐉", "sport": "🏅",
}

def normalise_famille(categorie):
    """Deduit le slug de famille (pour la couleur, l'emoji par defaut et les
    filtres de partie) a partir de la colonne categorie. Le cinema et la
    television sont sortis de "Culture" (trop gros, et assez distincts pour
    etre filtres separement) ; "Jeu video & ..." n'a pas de "/" donc se
    detecte par prefixe avant de retomber sur la regle generale (mot avant
    le premier "/")."""
    if categorie.startswith("Jeu vidéo") or categorie.startswith("Jeu video"):
        return "jeuxvideo"
    if categorie.startswith("Culture / Cinéma") or categorie.startswith("Culture / Cinema"):
        return "cinema"
    if categorie.startswith("Culture / Télévision") or categorie.startswith("Culture / Television"):
        return "television"
    famille = categorie.split('/')[0].strip()
    return FAMILLE_SLUG.get(famille, "culture")

def formatte_date_js(valeur):
    """Ecrit un literal JS propre pour la date. Au-dela de 10^15, un entier
    Python exact (issu de int(float)) produit une chaine de dizaines de
    chiffres illisible dans cartes.js : on ecrit alors une notation
    exponentielle (ex: 1e+30), tout aussi valide en JS et bien plus lisible.
    Les echelles cosmologiques (evaporation des trous noirs, etc.) n'ont de
    toute facon besoin que de l'ordre de grandeur, pas d'une valeur exacte."""
    v = float(valeur)
    if abs(v) >= 1e15:
        return repr(v)
    return str(int(v))

FIAB_SLUG = {
    "Avéré": "avere",
    "Débattu par les historiens": "debattu",
    "Légende populaire": "legende",
}

wb = load_workbook(XLSX, data_only=True)
ws_cartes = wb['Cartes']
ws_liens = wb['Liens']

# --- Liens : regroupe par id ---
liens_par_id = {}
for row in ws_liens.iter_rows(min_row=2, values_only=True):
    cid, type_, label, url = row[0], row[1], row[2], row[3]
    if cid is None:
        continue
    liens_par_id.setdefault(int(cid), []).append({"type": type_, "label": label, "url": url})

# --- Cartes ---
headers = [c.value for c in ws_cartes[1]]
col_idx = {h: i for i, h in enumerate(headers) if h}

cartes = []
for row in ws_cartes.iter_rows(min_row=2, values_only=True):
    cid = row[col_idx['id']]
    if cid is None:
        continue
    cid = int(cid)
    categorie = row[col_idx['categorie']]
    slug = normalise_famille(categorie)
    fiabilite = row[col_idx['fiabilite']]
    image = row[col_idx['image']]

    cartes.append({
        "id": cid,
        "categorie": categorie,
        "famille": slug,
        "titre": row[col_idx['titre']],
        "date": row[col_idx['date']],
        "emoji": EMOJIS.get(cid, EMOJI_PAR_FAMILLE.get(slug, "🃏")),
        "image": (image or "").strip(),
        "description_courte": row[col_idx['description_courte']],
        "description_longue": row[col_idx['description_longue']],
        "anecdote": row[col_idx['anecdote']] or "",
        "fiabilite": FIAB_SLUG.get(fiabilite, "avere"),
        "liens": liens_par_id.get(cid, []),
    })

cartes.sort(key=lambda c: c["id"])
ids = [c["id"] for c in cartes]
assert len(ids) == len(set(ids)), "des ids sont en double dans la feuille Cartes"
attendu = list(range(1, len(cartes) + 1))
if ids != attendu:
    manquants = sorted(set(attendu) - set(ids))
    print(f"ATTENTION : ids non contigus (il manque {manquants}) -- verifie la colonne id du xlsx.")

print("Total cartes generees:", len(cartes))
dates = [c["date"] for c in cartes]
print("Dates min/max:", min(dates), max(dates))
avec_image = sum(1 for c in cartes if c["image"])
print("Cartes avec image renseignee:", avec_image, "/", len(cartes))

# --- Ecriture du JS ---
lines = []
lines.append("// Genere automatiquement a partir de data/cartes_maitre.xlsx (feuilles Cartes + Liens)")
lines.append("// Ne pas editer ce fichier a la main : modifier le xlsx puis lancer generer_cartes.py")
lines.append("// image: nom de fichier dans images/ (vide = emoji utilise a la place)")
lines.append("// liens: tableau de { type: 'youtube'|'wikipedia'|'publication'|'livre'|'autre', label, url }")
lines.append("const BASE_CARTES = [")
for c in cartes:
    lines.append("  {")
    lines.append(f'    id: {c["id"]},')
    lines.append(f'    categorie: {json.dumps(c["categorie"], ensure_ascii=False)},')
    lines.append(f'    famille: {json.dumps(c["famille"], ensure_ascii=False)},')
    lines.append(f'    titre: {json.dumps(c["titre"], ensure_ascii=False)},')
    lines.append(f'    date: {formatte_date_js(c["date"])},')
    lines.append(f'    emoji: {json.dumps(c["emoji"], ensure_ascii=False)},')
    lines.append(f'    image: {json.dumps(c["image"], ensure_ascii=False)},')
    lines.append(f'    description_courte: {json.dumps(c["description_courte"], ensure_ascii=False)},')
    lines.append(f'    description_longue: {json.dumps(c["description_longue"], ensure_ascii=False)},')
    lines.append(f'    anecdote: {json.dumps(c["anecdote"], ensure_ascii=False)},')
    lines.append(f'    fiabilite: {json.dumps(c["fiabilite"], ensure_ascii=False)},')
    lines.append(f'    liens: {json.dumps(c["liens"], ensure_ascii=False)}')
    lines.append("  },")
lines.append("];")
lines.append("")

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print("Ecrit:", OUT)
